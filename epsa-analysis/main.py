# -*- coding: utf-8 -*-
# ----------------------------------------------------------------
# main: Main script from which all modules are run
# ----------------------------------------------------------------
# Created by    : Ryan French
# Created date  : 07/19/2023
# ----------------------------------------------------------------

from utilities import *
from analysis_resistor import *
from analysis_capacitor import *
from analysis_inductor import *
from prettify_excel import *

import sys
import os.path
import pint
from configparser import ConfigParser
from dataclasses import dataclass
from typing import List
import csv
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl_image_loader import SheetImageLoader

# Set up exception hook
sys.excepthook = EPSA_except_hook

def create_xlsx(filename=EXCEL_FILENAME):
    """Create the xlsx file if it doesn't exist

    Parameters
    ----------
    filename : string
        Filename of the Excel workbook. Must contain the .xlsx extension.

    Returns
    ----------
    None
    """
    print(f"Creating Excel file {filename}...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    wb.save(filename)
    wb.close()

def fill_in_overview(filename=EXCEL_FILENAME, pass_lists=[]):
    """Create the xlsx file if it doesn't exist

    Parameters
    ----------
    filename : string
        Filename of the Excel workbook. Must contain the .xlsx extension.
    pass_lists : list(Passes class)
        List of the Passes class to track how many passes each class of EPSA component has

    Returns
    ----------
    None
    """
    print(f"Styling Overview sheet...")

    wb = load_workbook(filename)
    ws = wb.active
    # Fill in static texts and styles
    style_overview_sheet(ws)

    # Fill in pass overview table
    for pass_list in pass_lists:
        if pass_list.name == "Resistors":
            r_passes = pass_list.pass_list[0].pass_num
            v_passes = pass_list.pass_list[1].pass_num
            p_passes = pass_list.pass_list[2].pass_num
            t_passes = pass_list.pass_list[3].pass_num
            ws["B10"] = "Resistors"
            ws["B10"].font = Font(size=12)
            ws["B10"].alignment = Alignment(horizontal="center", vertical="center") 
            ws["C10"] = pass_list.total
            ws["C10"].font = Font(size=12)
            ws["C10"].alignment = Alignment(horizontal="center", vertical="center") 
            # Overall Passes
            ws["D10"] = "Overall"
            ws["D10"].font = Font(size=12)
            ws["D10"].alignment = Alignment(horizontal="center", vertical="center") 
            ws["E10"] = r_passes
            ws["E10"].font = Font(size=12)
            ws["E10"].alignment = Alignment(horizontal="center", vertical="center") 
            # V Passes
            ws["D11"] = "V Pass"
            ws["D11"].font = Font(size=12)
            ws["D11"].alignment = Alignment(horizontal="center", vertical="center") 
            ws["E11"] = v_passes
            ws["E11"].font = Font(size=12)
            ws["E11"].alignment = Alignment(horizontal="center", vertical="center") 
            # P Passes
            ws["D12"] = "P Pass"
            ws["D12"].font = Font(size=12)
            ws["D12"].alignment = Alignment(horizontal="center", vertical="center") 
            ws["E12"] = p_passes
            ws["E12"].font = Font(size=12)
            ws["E12"].alignment = Alignment(horizontal="center", vertical="center") 
            # T Passes
            ws["D13"] = "T Pass"
            ws["D13"].font = Font(size=12)
            ws["D13"].alignment = Alignment(horizontal="center", vertical="center") 
            ws["E13"] = t_passes
            ws["E13"].font = Font(size=12)
            ws["E13"].alignment = Alignment(horizontal="center", vertical="center") 
        
        elif pass_list.name == "Capacitors":
            c_passes = pass_list.pass_list[0].pass_num
            v_passes = pass_list.pass_list[1].pass_num
            t_passes = pass_list.pass_list[2].pass_num
            ws["B14"] = "Capacitors"
            ws["B14"].font = Font(size=12)
            ws["B14"].alignment = Alignment(horizontal="center", vertical="center") 
            ws["C14"] = pass_list.total
            ws["C14"].font = Font(size=12)
            ws["C14"].alignment = Alignment(horizontal="center", vertical="center") 
            # Overall Passes
            ws["D14"] = "Overall"
            ws["D14"].font = Font(size=12)
            ws["D14"].alignment = Alignment(horizontal="center", vertical="center") 
            ws["E14"] = c_passes
            ws["E14"].font = Font(size=12)
            ws["E14"].alignment = Alignment(horizontal="center", vertical="center") 
            # V Passes
            ws["D15"] = "V Pass"
            ws["D15"].font = Font(size=12)
            ws["D15"].alignment = Alignment(horizontal="center", vertical="center") 
            ws["E15"] = v_passes
            ws["E15"].font = Font(size=12)
            ws["E15"].alignment = Alignment(horizontal="center", vertical="center") 
            # T Passes
            ws["D16"] = "T Pass"
            ws["D16"].font = Font(size=12)
            ws["D16"].alignment = Alignment(horizontal="center", vertical="center") 
            ws["E16"] = t_passes
            ws["E16"].font = Font(size=12)
            ws["E16"].alignment = Alignment(horizontal="center", vertical="center") 

        elif pass_list.name == "Inductors":
            l_passes = pass_list.pass_list[0].pass_num
            v_passes = pass_list.pass_list[1].pass_num
            t_passes = pass_list.pass_list[2].pass_num
            ws["B17"] = "Inductors"
            ws["B17"].font = Font(size=12)
            ws["B17"].alignment = Alignment(horizontal="center", vertical="center") 
            ws["C17"] = pass_list.total
            ws["C17"].font = Font(size=12)
            ws["C17"].alignment = Alignment(horizontal="center", vertical="center") 
            # Overall Passes
            ws["D17"] = "Overall"
            ws["D17"].font = Font(size=12)
            ws["D17"].alignment = Alignment(horizontal="center", vertical="center") 
            ws["E17"] = l_passes
            ws["E17"].font = Font(size=12)
            ws["E17"].alignment = Alignment(horizontal="center", vertical="center") 
            # V Passes
            ws["D18"] = "V Pass"
            ws["D18"].font = Font(size=12)
            ws["D18"].alignment = Alignment(horizontal="center", vertical="center") 
            ws["E18"] = v_passes
            ws["E18"].font = Font(size=12)
            ws["E18"].alignment = Alignment(horizontal="center", vertical="center") 
            # T Passes
            ws["D19"] = "T Pass"
            ws["D19"].font = Font(size=12)
            ws["D19"].alignment = Alignment(horizontal="center", vertical="center") 
            ws["E19"] = t_passes
            ws["E19"].font = Font(size=12)
            ws["E19"].alignment = Alignment(horizontal="center", vertical="center") 


    wb.save(filename)
    wb.close()


def main():
    print(f"EPSA Analysis Script v. {SCRIPT_VERSION}")
    print(f"-----------------------------\n")

    # Create new Excel workbook
    create_xlsx(EXCEL_FILENAME)

    # Resistors
    analyze_resistors()
    write_resistors_to_excel(EXCEL_FILENAME)
    r_passes = get_R_passes()

    # Capacitors
    analyze_capacitors()
    write_capacitors_to_excel(EXCEL_FILENAME)
    c_passes = get_C_passes()

    # Inductors 
    analyze_inductors()
    write_inductors_to_excel(EXCEL_FILENAME)
    l_passes = get_L_passes()

    # Excel styling 
    style_epsa_sheets(EXCEL_FILENAME)
    auto_column_width(EXCEL_FILENAME)
    highlight_passes(EXCEL_FILENAME)

    fill_in_overview(EXCEL_FILENAME, [
        r_passes,
        c_passes,
        l_passes
    ])

    input("\n\nAnalysis done. Press Enter to exit.")

if __name__ == "__main__":
    main()